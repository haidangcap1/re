
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import random
import platform
import sys
import os
import json
import pickle
import click
import openpyxl
from openpyxl import load_workbook
import random
from fake_useragent import UserAgent
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import WebDriverException, TimeoutException, MoveTargetOutOfBoundsException
from selenium.common.exceptions import TimeoutException
from pathlib import Path
from webdriver_manager import driver
from webdriver_manager.chrome import ChromeDriverManager


wb = load_workbook('a.xlsx')
sh = wb.active
maxrow = sh.max_row
wb = openpyxl.load_workbook('a.xlsx')
sheet = wb['Sheet1']
a = random.randint(1,maxrow)
tensp = sheet.cell(row=a, column=1)
linksp = sheet.cell(row=a, column=2)
d = tensp.value 
dd = linksp.value
z = 1
j = 1
p = random.randint(1,3)
BASE_PATH = Path('__file__').resolve().parent
DRIVERS_PATH = BASE_PATH / "drivers"
os_version = platform.system()

class Linkedin:
    def __init__(self, port=None):
        self.port = port

    def create_driver(self, headless):
        options = Options()
        if self.port:
            options.add_experimental_option(
                "debuggerAddress", f"localhost:{self.port}")
        else:
            ua = UserAgent()
            assd = ua.random
            user_agent = ua.random
            options.add_argument(f'user-agent={user_agent}')
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option(
                "excludeSwitches", ["enable-automation"])
            
            options.add_argument('--log-level=3')
            prefs = {"profile.default_content_setting_values.notifications": 2}
            options.add_experimental_option("prefs", prefs)
            options.headless = headless

        if os_version == "Linux":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_linux")
        elif os_version == "Darwin":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_mac")
        elif os_version == "Windows":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_windows.exe")
        else:
            driver = webdriver.Chrome(
                ChromeDriverManager().install(), options=options)
            return driver
        driver = webdriver.Chrome(executable_path=executable, options=options)
        return driver
    
    def  khoitao(self , driver , a):
        try:
            
            driver.get('https://www.shopee.vn')
            time.sleep(random.randint(3,7))
            driver.find_element_by_xpath('//*[@id="modal"]/div/div/div[2]/div').click()
        except TimeoutException:
          driver.close()
    def tukhoa(self,driver,b):
        try:
            
            time.sleep(random.randint(5,10))
            driver.execute_script('window.scrollTo(0,1000)')
            driver.find_element_by_xpath("//input").send_keys(b)
            time.sleep(random.randint(1,4))
            driver.find_element_by_xpath("//input").send_keys(Keys.ENTER)
        except TimeoutException:
          driver.close()

    def tuongtacsanpham1(self,driver,c):
        
        try:
           
            driver.execute_script('window.scrollTo(0,1500)')
            time.sleep(random.randint(5,10))
            driver.execute_script('window.scrollTo(0,2300)')
            time.sleep(random.randint(5,10))
            next1 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//div[contains(@class, '_10Wbs- _5SSWfi UjjMrh')]")))
            driver.execute_script( "return arguments[0].scrollIntoView(true);", next1)
            driver.execute_script('window.scrollTo(0,3000)')
            
            next12 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//div[contains(@class, '_10Wbs- _5SSWfi UjjMrh')]")))
            driver.execute_script( "return arguments[0].scrollIntoView(true);", next12)
            driver.execute_script('window.scrollTo(0,1000)')
            time.sleep(random.randint(5,10))

            
        except TimeoutException:
          driver.close()
        
    def chuyentrang(self,driver, ll):
        
    
        try:
            
            next = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//button[contains(@class, 'shopee-icon-button shopee-icon-button--right')]")))
            driver.execute_script( "return arguments[0].scrollIntoView(true);", next)
            ActionChains(driver).move_to_element(next).click().perform()
                
        except TimeoutException:
          driver.close()
    def bonus(self,driver,sd):
        driver.get(linksp.value)
    def tuongtacsanpham(self,driver, assc):
       
        try:
            
            time.sleep(random.randint(5,10))
            driver.execute_script('window.scrollTo(0,2000)')
           
            submit_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class = '_3-_YTZ']")))
            ActionChains(driver).move_to_element(submit_button).click().perform()
            i = random.randint(1,5)
            while (i <= 8):
                time.sleep(random.randint(10,25))
                ads= WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class = '_2cmzNg']")))
                ActionChains(driver).move_to_element(ads).click().perform()
                i = i + 1
            
            driver.refresh()
            vcl = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                            (By.XPATH, "//div[contains(@class, 'OitLssRu')]")))
            ActionChains(driver).move_to_element(vcl).click().perform()

            time.sleep(random.randint(20,25))
            driver.execute_script("window.scrollTo(0, 3000);")
            time.sleep(random.randint(20,25))
            driver.execute_script("window.scrollTo(0, 4000);")
            g = random.randint(1,3)
            try:
                while(g <= 5):
                    vcl1 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//div[contains(@class, 'product-rating-overview__filter')]")))
                    time.sleep(random.randint(10,15)) 
                    driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl1)
                    time.sleep(random.randint(10,15))
                    ActionChains(driver).move_to_element(vcl1).click().perform()
                    driver.execute_script('window.scrollTo(0,2000)')
                    g = g + 1 
                    time.sleep(random.randint(10,15))
            except TimeoutException:
                pass  
            try:
                cc=  random.randint(1,5)
                while(cc <= 5):
                    vcl12 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//button[contains(@class, 'shopee-icon-button shopee-icon-button--right')]")))
                    time.sleep(random.randint(10,15)) 
                    driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl12)
                    time.sleep(random.randint(10,15))
                    ActionChains(driver).move_to_element(vcl12).click().perform()
                    cc = cc + 1
            except TimeoutException:
                pass
        except TimeoutException:
             driver.close()
    def sanphamtuongtu(qq):
         try:
            vcl123 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//div[contains(@class, 'carousel-arrow carousel-arrow--next carousel-arrow--hint')]")))
        
            driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl123)
          
            ActionChains(driver).move_to_element(vcl123).click().perform()
            time.sleep(random.randint(5,10))
            vcl124 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//div[contains(@class, '_10Wbs- UjjMrh')]")))
            
            driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl124)
            
            ActionChains(driver).move_to_element(vcl124).click().perform()
         except TimeoutException:
             pass
    

@click.command()
@click.option("-h", "--headless", is_flag=True, help="Hide the borwser's window.")

def main(headless):
    try:
         ln = Linkedin()
         driver12 = ln.create_driver('a')
         ln.khoitao(driver12, 'a')
         ln.tukhoa(driver12,d)
         ii  = random.randint(1,3)
         while ( ii <= 5):
            ln.tuongtacsanpham1(driver12, 's')
            time.sleep(random.randint(12,15))
            ln.chuyentrang(driver12,'as')
            ii =  ii + 1
         iii = random.randint(1,3)
         ln.bonus(driver12,'sd')
         while (iii <= 6):
             ln.tuongtacsanpham(driver12,'d')
             time.sleep(random.randint(10,23))
             ln.chuyentrang(driver12,'q')
             iii = iii + 1
         ln.close() 
    except TimeoutException:
        
        ln.close()
      
if __name__ == "__main__":
    main()  


